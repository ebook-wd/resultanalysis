#!/usr/bin/env python

"""
resultanalysis.py: Download the students result from website http://www.tekerala.org/.

Usage: Type at the command prompt

python copytxt.py -i <inputfile> -o <outputfile> 

inputfile     =  name of the input file. Must be a excel file  (with extension .xlsx)  containing 
                register number in first column and date of birth in second coulmn.
                Register no and DOB must be in the first excel sheet with name Sheet1 
outputfile    =  name of the output file. Must be a .xlsx file.

eg:

python resultanalysis.py -i regno.xlsx -o result.xlsx

Bugs and issues can be repoted to author at williamdoyleaf@gmail.com
"""

__author__      = "William Doyle A F"
__copyright__   = "Copyright 2014"
__license__ 	= "GNU GPLv3 (http://www.gnu.de/documents/gpl-3.0.en.html)"
__version__ 	= "5.0"
__maintainer__  = "William Doyle A F"
__email__       = "williamdoyleaf@gmail.com"


try:
    import sys
    import os
    import getopt
except:
    sys.exit("Import error:0")
try:
    import requests
except:
    print "Module requests is not available. Install it  using command \n"
    print " \"sudo apt-get install python-requests\" \n"
    sys.exit("Import error:1")
try:
    from bs4 import BeautifulSoup
except:
    print "Module bs4 is not available. Install it  using command \n"
    print " \"sudo apt-get install python-bs4\" \n"
    sys.exit("Import error:2")
try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
except:
    print "Module openpyxl is not available. Install it  using command \n"
    print " \"sudo apt-get install python-openpyxl\" \n"
    sys.exit("Import error:3")

URL = 'http://www.tekerala.org/student_detailss/index.php'

def main():
    Input_File_Name  = ''
	# Assigns the out put file name
    Output_File_Name = ''

    try:
      		opts, args = getopt.getopt(sys.argv[1:],"hi:o:",["infile=","outfile="])
    except getopt.GetoptError:
      		print 'copytxt.py -i <inputfile> -o <outputfile> '
      		sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            usage()
            sys.exit()
        elif opt in ("-i", "--infile"):
            Input_File_Name  = arg
        elif opt in ("-o", "--outfile"):
            Output_File_Name = arg  
    		
	wb_out = Workbook()
    wb_out.save(Output_File_Name)
    wb_out = load_workbook(filename= Output_File_Name)
    ws_out = wb_out.create_sheet("Sheet1", 0)	

    s = requests.session()
  
    wb = load_workbook(Input_File_Name)
    sheet = wb.get_sheet_by_name('Sheet1')
    ws_row = 1
    ws_col = 1 
    set=1
    i = 0
    for r in  range(1,sheet.max_row+1):
        i = i + 1
        regno = str(sheet.cell(row=i, column=1).value)
        dob = str(sheet.cell(row=i, column=2).value)
        
        login_data = {
            'regno':regno,
            'dob':dob,
            'captcha':'captcha'        
        }

        r = s.post(URL, data=login_data)
        r = s.get('http://www.tekerala.org/student_detailss/mark_list_view.php')
        html=r.text
        soup = BeautifulSoup(html,"lxml")
        td = soup.findAll('td',width="119", valign="top")
        for t in td:            
            rn = t.findAll('strong')
            rno = rn[0].string               
        if rno == regno :
            print(regno +" Fetching result....")
        else:
            print("Register or DOB is not mathching")
            ws_row = ws_row + 1
            ws_col = 1
            ws_out.cell(row= ws_row , column=ws_col).value = regno
            continue
        
        # Try accessing a page that requires you to be logged in
        r = s.get('http://www.tekerala.org/student_detailss/rslt_mark_view.php?year_exam=2016&mon_exam=Apr&semester=2')
        html=r.text
        soup = BeautifulSoup(html,"lxml")
        table = soup.find("table", border=0, cellpadding=4, cellspacing=1)
        if set==1:
            ws_out.cell(row= ws_row , column=ws_col).value = "RegNo"
            ws_col = ws_col +1
            for row in table.findAll('tr')[1:]:
                col = row.findAll('td')
                sub = col[1].string
                record = (sub)     
                ws_out.cell(row= ws_row , column=ws_col).value = record
                ws_col = ws_col +1
        set=0   
        ws_row = ws_row + 1
        ws_col = 1
        ws_out.cell(row= ws_row , column=ws_col).value = regno
        ws_col = ws_col +1
        for row in table.findAll('tr')[1:]:
            col = row.findAll('td')
            grade = col[4].string
            record = (grade)
            ws_out.cell(row= ws_row , column=ws_col).value = record
            ws_col = ws_col +1           
    wb_out.save(Output_File_Name)
def usage():
 print """	
	Usage: Type at the command prompt
	python copytxt.py -i <inputfile> -o <outputfile> 
    inputfile     =  name of the input file. Must be a excel file (with extension .xlsx) containing 
                        register number in first column and date of birth in second coulmn.
                    Register no and DOB must be in the first excel sheet with name Sheet1 
    outputfile    =  name of the output file. Must be a .xlsx file.

    eg:

    python resultanalysis.py -i regno.xlsx -o result.xlsx
	"""	

if __name__ == '__main__':
    main()
